from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.cache import never_cache

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.http import HttpResponse
from reportlab.pdfgen import canvas  # Añadida esta importación
from django.contrib.staticfiles import finders
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from inicio.forms import ValoracionForm, UserForm
from inicio.models import UserProfile, Valoracion
import inicio.views as traer

import logging

logger = logging.getLogger(__name__)

@never_cache
@login_required(login_url='acceso_denegado')
def fetch_user_details(request):
    logger.info('fetch_user_details view called')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        documento = request.GET.get('documento')
        logger.info(f'Documento recibido: {documento}')
        user = UserProfile.objects.filter(documento=documento).first()
        if user:
            data = {
                'tipo': user.tipo,
                'nombre': user.nombre,
                'correo': user.correo,
                'direccion': user.direccion,
                'edad': user.edad,
                'ocupacion': user.ocupacion,
                'celular': user.celular,
                'acudiente': user.acudiente,
            }
            return JsonResponse(data)
    return JsonResponse({}, status=404)

@never_cache
@login_required(login_url='acceso_denegado')
def listhistorias(request):
    # Si es superusuario, obtiene todas las historias
    if request.user.is_superuser:
        historias = Valoracion.objects.all()
    else:
        # Si no es superusuario, solo obtiene las historias del usuario actual
        historias = Valoracion.objects.filter(user=request.user)

    # Filtro por fecha
    fecha = request.GET.get('fecha')
    if fecha:
        historias = historias.filter(fecha_historia=fecha)

    return render(request, 'listhistorias.html', {
        'historias': historias
    })

@never_cache
@login_required(login_url='acceso_denegado')
def verhistorias(request, id):
    historia = Valoracion.objects.select_related('user').get(id=id)
    return render(request, 'verhistorias.html', {'historia': historia})

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def eliminarhistorias(request, id):
    historiaseliminar = get_object_or_404(Valoracion, id=id)
    
    if request.method == 'POST':
        historiaseliminar.delete()
        return redirect('listhistorias')
    
    return redirect('listhistorias')

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def crearhistorias(request):
    if request.method == 'POST':
        documento = request.POST.get('documento')
        user = UserProfile.objects.filter(documento=documento).first()
        
        formularioI = UserForm(request.POST)
        formularioII = ValoracionForm(request.POST)
        
        if user:
            formularioI = UserForm(request.POST, instance=user)
            if formularioI.is_valid() and formularioII.is_valid():
                user = formularioI.save()
                valoracion = formularioII.save(commit=False)
                valoracion.user = user
                valoracion.save()
                messages.success(request, 'Historia clínica creada con éxito.')
                return redirect('listhistorias')
            else:
                print("Errores en formularioI:", formularioI.errors)
                print("Errores en formularioII:", formularioII.errors)
        else:
            messages.error(request, 'El usuario no existe. Por favor, verifique el número ingresado.')
    else:
        formularioI = UserForm()
        formularioII = ValoracionForm()
    
    context = {
        'formularioI': formularioI,
        'formularioII': formularioII
    }
    return render(request, 'crearhistorias.html', context)

from openpyxl import Workbook

def reporte_historia_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historias Clínicas"

    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 55
    img.width = 75
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:D1')
    ws['B1'] = "LABORATORIO DENTAL"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = "Historias Clínicas"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Paciente", "Documento", "Fecha Historia", "Tratamiento Medicación", "Reacciones Alérgicas", 
               "Trastorno Tensión Arterial", "Diabetes", "Trastornos Emocionales"]
    ws.append(headers)

    # Cambiamos el color de fondo del encabezado a #cab97d
    header_fill = PatternFill(start_color="cab97d", end_color="cab97d", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    historias = Valoracion.objects.select_related('user').all()
    for historia in historias:
        ws.append([
            historia.user.nombre,
            historia.documento,
            historia.fecha_historia.strftime('%d-%m-%Y') if historia.fecha_historia else 'N/A',
            dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.tratamiento_medicacion, 'N/A'),
            dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.reacciones_alergicas, 'N/A'),
            dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.transtorno_tension_arterial, 'N/A'),
            dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.diabetes, 'N/A'),
            dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.transtornos_emocionales, 'N/A'),
        ])

    column_widths = [20, 15, 20, 25, 25, 25, 15, 25]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Historias_Clinicas.xlsx"'
    
    wb.save(response)
    return response


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.logo_path = finders.find('img/logo.png')

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        page_width, page_height = landscape(letter)
        self.setFont("Helvetica", 9)
        self.drawRightString(page_width - 30, 30, f"Página {self._pageNumber} de {page_count}")
        
        self.saveState()
        img = ImageReader(self.logo_path)
        img_width = 4 * inch
        img_height = img_width * img.getSize()[1] / img.getSize()[0]
        
        # Calcular posición para centrar la imagen
        x = (page_width - img_width) / 2
        y = (page_height - img_height) / 2
        
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.1))  # 10% opacity
        self.drawImage(img, x, y, width=img_width, height=img_height, mask='auto')
        self.restoreState()

def reporte_historia_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=1))

    elements.append(Paragraph("LABORATORIO DENTAL", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Historias Clínicas", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Definir estilos para el encabezado y el contenido de la tabla
    estilo_encabezado = ParagraphStyle(
        'EstiloEncabezado',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        spaceAfter=0,
        leading=10,
    )
    estilo_contenido = ParagraphStyle(
        'EstiloContenido',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        spaceAfter=0,
        leading=10,
    )

    # Crear Paragraphs para el encabezado
    encabezados = [
        Paragraph("Paciente", estilo_encabezado),
        Paragraph("Documento", estilo_encabezado),
        Paragraph("Fecha Historia", estilo_encabezado),
        Paragraph("Tratamiento Medicación", estilo_encabezado),
        Paragraph("Reacciones Alérgicas", estilo_encabezado),
        Paragraph("Trastorno Tensión Arterial", estilo_encabezado),
        Paragraph("Diabetes", estilo_encabezado),
        Paragraph("Trastornos Emocionales", estilo_encabezado)
    ]

    data = [encabezados]

    historias = Valoracion.objects.select_related('user').all()
    for historia in historias:
        fila = [
            Paragraph(historia.user.nombre, estilo_contenido),
            Paragraph(str(historia.documento), estilo_contenido),
            Paragraph(historia.fecha_historia.strftime('%d-%m-%Y') if historia.fecha_historia else 'N/A', estilo_contenido),
            Paragraph(dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.tratamiento_medicacion, 'N/A'), estilo_contenido),
            Paragraph(dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.reacciones_alergicas, 'N/A'), estilo_contenido),
            Paragraph(dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.transtorno_tension_arterial, 'N/A'), estilo_contenido),
            Paragraph(dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.diabetes, 'N/A'), estilo_contenido),
            Paragraph(dict(Valoracion.OPCIONES_SI_NO_NO_SABE).get(historia.transtornos_emocionales, 'N/A'), estilo_contenido),
        ]
        data.append(fila)

    table = Table(data, colWidths=[110, 70, 80, 90, 90, 95, 70, 95])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cab97d")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)

    elements.append(table)

    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Historias_Clinicas.pdf"'
    
    return response



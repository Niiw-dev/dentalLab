import json
import logging

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache
from django.contrib.staticfiles import finders

from google.auth.transport.requests import Request

from inicio.forms import UserForm
from inicio.models import UserProfile
from inicio import views as traer

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph


@never_cache
@login_required(login_url='acceso_denegado')
def fetch_user_details(request):
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        documento = request.GET.get('documento')
        print(f'Documento recibido: {documento}')  # Agrega esta línea
        user = UserProfile.objects.filter(documento=documento).first()
        if user:
            data = {
                'tipo': user.tipo,
                'nombre': user.nombre,
                'correo': user.correo,
                'direccion': user.direccion,
                'edad': user.edad,
                'ocupacion': user.ocupacion,
                'celular': user.celular,
                'acudiente': user.acudiente,
            }
            return JsonResponse(data)
    return JsonResponse({}, status=404)


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def crearcuentas(request):
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('listcuentas')  
    else:
        form = UserForm()
    
    return render(request, 'crearcuentas.html', {'form': form})


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def listcuentas(request):
    cuentas = UserProfile.objects.all()
    return render(request, 'listcuentas.html', {'cuentas': cuentas})


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def reactivar_cuenta(request, id):
    cuenta = get_object_or_404(UserProfile, id=id)
    if request.method == 'POST':
        cuenta.is_active = True
        cuenta.save()
        messages.success(request, 'La cuenta ha sido reactivada exitosamente.')
    return redirect('listcuentas')


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def editarcuentas(request, id):
    cuenta = get_object_or_404(UserProfile, id=id)
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=cuenta)
        if form.is_valid():
            form.save()
            return redirect('listcuentas')
    else:
        form = UserForm(instance=cuenta)
    return render(request, 'editarcuentas.html', {'form': form})


@never_cache
@login_required(login_url='acceso_denegado')
def cambiar_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Tu contraseña ha sido actualizada con éxito!')
            return redirect('dashboard')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'cambiar_password.html', {'form': form})


@never_cache
@login_required(login_url='acceso_denegado')
def eliminar_cuenta(request, id):
    cuenta = get_object_or_404(UserProfile, id=id)

    # Verifica si el usuario actual es el dueño de la cuenta o un superusuario
    if request.user.id != cuenta.id and not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'No tienes permiso para desactivar esta cuenta.'})

    cuenta.is_active = False
    cuenta.save()
    
    # Si el usuario está desactivando su propia cuenta, cierra la sesión
    if request.user.id == cuenta.id:
        logout(request)
        return JsonResponse({'success': True, 'message': 'Tu cuenta ha sido desactivada y has sido desconectado.', 'redirect': '/loginregister/'})
    else:
        return JsonResponse({'success': True, 'message': 'La cuenta ha sido desactivada.'})


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def activar_cuenta(request, id):
    cuenta = get_object_or_404(UserProfile, id=id)

    # Verifica si el usuario actual es un superusuario
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': 'No tienes permiso para activar esta cuenta.'})

    # Cambia el estado de is_active a True
    cuenta.is_active = True
    cuenta.save()

    return JsonResponse({'success': True, 'message': 'La cuenta ha sido activada exitosamente.'})


def reporte_pacientes_excel(request):
    data = json.loads(request.body)
    estado_filtro = data.get('estado', None)
    tipo_documento_filtro = data.get('tipoDoc', None)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    logger.debug(f"Estado filtro: {estado_filtro}")
    logger.debug(f"Documento filtro: {tipo_documento_filtro}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 21
    img.width = 39
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:H1')
    ws['B1'] = "LABORATORIO DENTAL"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = "Pacientes"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Tipo", "Documento", "Nombre", "Correo", "Dirección", "Celular", "Ocupación", "Edad"]
    ws.append(headers)

    header_fill = PatternFill(start_color="cab97d", end_color="cab97d", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    pacientes = UserProfile.objects.all()

    if estado_filtro:
        pacientes = pacientes.filter(is_active=(0 if estado_filtro == 'inactivo' else 1))

    if tipo_documento_filtro:
        pacientes = pacientes.filter(tipo=tipo_documento_filtro)

    for paciente in pacientes:
        tipo = dict(paciente.TIPO_CHOICES).get(paciente.tipo, "N/A")
        documento = paciente.documento
        nombre = paciente.nombre
        correo = paciente.correo
        direccion = paciente.direccion or "N/A"
        celular = paciente.celular or "N/A"
        ocupacion = paciente.ocupacion or "N/A"
        edad = paciente.edad or "N/A"

        ws.append([
            tipo,
            documento,
            nombre,
            correo,
            direccion,
            celular,
            ocupacion,
            edad
        ])

    column_widths = [15, 20, 25, 30, 30, 15, 24, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_pacientes.xlsx"'

    wb.save(response)
    return response


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.logo_path = finders.find('img/logo.png')

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(letter[0] - 30, 30, f"Página {self._pageNumber} de {page_count}")
        
        self.saveState()
        img = ImageReader(self.logo_path)
        img_width = 4 * inch
        img_height = img_width * img.getSize()[1] / img.getSize()[0]
        
        x = (letter[0] - img_width) / 2
        y = (letter[1] - img_height) / 2
        
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.1))
        self.drawImage(img, x, y, width=img_width, height=img_height, mask='auto')
        self.restoreState()


def reporte_pacientes_pdf(request):
    data = json.loads(request.body)
    estado_filtro = data.get('estado', None)
    tipo_documento_filtro = data.get('tipoDoc', None)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    logger.debug(f"Estado filtro: {estado_filtro}")
    logger.debug(f"Documento filtro: {tipo_documento_filtro}")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=1))

    elements.append(Paragraph("LABORATORIO DENTAL", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Pacientes", styles['Heading2']))
    elements.append(Spacer(1, 12))

    data = [["Tipo", "Documento", "Nombre", "Correo", "Dirección", "Celular", "Ocupación", "Edad"]]

    pacientes = UserProfile.objects.all()

    if estado_filtro:
        pacientes = pacientes.filter(is_active=0 if estado_filtro == 'inactivo' else 1)

    if tipo_documento_filtro:
        pacientes = pacientes.filter(tipo=tipo_documento_filtro)

    for paciente in pacientes:
        tipo = dict(paciente.TIPO_CHOICES).get(paciente.tipo, "N/A")
        data.append([
            tipo,
            paciente.documento,
            paciente.nombre,
            paciente.correo,
            paciente.direccion or "N/A",
            paciente.celular or "N/A",
            paciente.ocupacion or "N/A",
            str(paciente.edad) if paciente.edad else "N/A"
        ])

    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cab97d")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    elements.append(table)

    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_pacientes.pdf"'

    return response


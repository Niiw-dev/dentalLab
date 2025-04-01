import json
import os
from datetime import datetime, timedelta
from io import BytesIO
import logging

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.staticfiles import finders
from django.core.files.storage import FileSystemStorage
from django.core.mail import send_mail
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET, require_POST

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from inicio import views as traer
from inicio.forms import CitaForm, UserForm
from inicio.models import Cita, Fecha, UserProfile

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


CLIENT_SECRETS_FILE = os.path.join(settings.BASE_DIR, 'config/credentials.json')

SCOPES = [
    'https://www.googleapis.com/auth/calendar',
    'https://www.googleapis.com/auth/calendar.readonly'
]

def get_google_calendar_service(request):
    """
    Obtiene el servicio de Google Calendar utilizando las credenciales almacenadas en la sesión del usuario.

    Args:
        request: El objeto de solicitud HTTP que contiene la sesión del usuario.

    Returns:
        Un objeto de servicio de Google Calendar si las credenciales son válidas.
        Redirige a la vista de inicio de OAuth si las credenciales no son válidas o no están disponibles.
    """
    if 'credentials' not in request.session:
        return redirect('initiate_oauth')
    
    credentials = Credentials(**request.session['credentials'])
    
    if not credentials.valid:
        if credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
            request.session['credentials'] = {
                'token': credentials.token,
                'refresh_token': credentials.refresh_token,
                'token_uri': credentials.token_uri,
                'client_id': credentials.client_id,
                'client_secret': credentials.client_secret,
                'scopes': credentials.scopes
            }
        else:
            return redirect('initiate_oauth')

    return build('calendar', 'v3', credentials=credentials)


def initiate_oauth(request):
    redirect_uri = request.build_absolute_uri(reverse('oauth2callback'))
    print(f"Redirect URI: {redirect_uri}")
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        redirect_uri=redirect_uri
    )
    authorization_url, state = flow.authorization_url(
        access_type='offline',
        include_granted_scopes='true',
    )
    request.session['state'] = state
    print(f"Authorization URL: {authorization_url}")
    return redirect(authorization_url)


def oauth2callback(request):
    
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    os.environ['OAUTHLIB_RELAX_TOKEN_SCOPE'] = '1'

    redirect_uri = request.build_absolute_uri(reverse('oauth2callback'))
    print(f"Redirect URI in oauth2callback: {redirect_uri}")
    print(f"Full request URL: {request.build_absolute_uri()}")
    
    state = request.session.get('state')
    if not state:
        return redirect('initiate_oauth')
    
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        state=state,
        redirect_uri=redirect_uri
    )
    
    try:
        flow.fetch_token(authorization_response=request.build_absolute_uri())
        credentials = flow.credentials
        
        request.session['credentials'] = {
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': credentials.scopes
        }
        
        messages.success(request, 'Autenticación con Google Calendar completada exitosamente.')
        return redirect('listcitas')
    
    except Warning as w:
        print(f"Advertencia durante la autenticación: {str(w)}")
        messages.warning(request, f'Se produjo una advertencia durante la autenticación: {str(w)}')
        return redirect('listcitas')
    
    except Exception as e:
        print(f"Error durante la autenticación: {str(e)}")
        messages.error(request, f'Error durante la autenticación: {str(e)}')
        return redirect('initiate_oauth')


@never_cache
@login_required(login_url='acceso_denegado')
@require_POST
def cancelar_cita(request, cita_id):
    try:
        cita = Cita.objects.get(id=cita_id)
        resultado = cita.cancelar_cita()
        if resultado:
            print("entró")
            service = get_google_calendar_service(request)
            if isinstance(service, HttpResponseRedirect):
                return service

            events_result = service.events().list(calendarId='primary', q=f'Cita ID: {cita_id}').execute()
            events = events_result.get('items', [])

            if events:
                event = events[0]
                service.events().delete(calendarId='primary', eventId=event['id']).execute()
                print(f"Evento de Google Calendar eliminado para la cita {cita_id}")
                
            recipient_list = [cita.paciente.correo, "facturacionldsg@gmail.com"]
            send_mail(
                'Cita Cancelada',
                f'Hola {cita.paciente.nombre},\n\nTu cita para el {cita.fecha_hora} ha sido cancelada .\nSi deseas agendar otra cita ingresa a nuestra plataforma. \n\n Saludos,\nTu Equipo de Citas \n Laboratorio Dental - Sandra Gavíria',
                settings.DEFAULT_FROM_EMAIL,
                recipient_list,
                fail_silently=False,
            )
            
            print(f"Cita {cita_id} cancelada. Nueva disponibilidad: {cita.fecha_hora.disponible}")
            return JsonResponse({'status': 'success'}, status=200)
        else:
            return JsonResponse({'status': 'error', 'message': 'No se pudo cancelar la cita'}, status=400)
    except Cita.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Cita no encontrada'}, status=404)
    except Exception as e:
        print(f"Error al cancelar cita: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@never_cache
@login_required(login_url='acceso_denegado')
def crearcitas(request):
    if request.method == 'POST':
        formulario = CitaForm(request.POST, user=request.user)
        if formulario.is_valid():
            fecha = formulario.cleaned_data['fecha']
            hora = formulario.cleaned_data['hora']

            if not fecha or not hora:
                messages.error(request, 'Por favor, selecciona tanto la fecha como la hora para la cita.')
                return redirect('listcitas')
            
            paciente = formulario.cleaned_data['paciente'] if request.user.is_superuser else request.user

            if Cita.objects.filter(paciente=paciente, estado='Programada').exists():
                messages.error(request, f'El paciente {paciente.nombre} ya tiene una cita programada.')
                return redirect('listcitas')

            cita = formulario.save(commit=False)
            cita.paciente = paciente
            cita.estado = 'Programada'

            print(f'Fecha: {fecha}, Hora: {hora}')

            fechas_existentes = Fecha.objects.filter(fecha=fecha, hora=hora)
            print(f"Fechas existentes para {fecha} {hora}: {fechas_existentes.count()} encontradas.")

            if fechas_existentes.count() > 1:
                messages.error(request, 'Se encontraron múltiples horarios. Por favor, selecciona otro.')
                return redirect('listcitas')
            elif fechas_existentes.exists():
                cita.fecha_hora = fechas_existentes.first()
            else:
                nueva_fecha = Fecha(fecha=fecha, hora=hora)
                nueva_fecha.save()
                cita.fecha_hora = nueva_fecha

            cita.save()

            recipient_list = [cita.paciente.correo, "facturacionldsg@gmail.com"]
            send_mail(
                'Recordatorio de Cita Programada',
                f'Hola {cita.paciente.nombre},\n\nTu cita ha sido programada para el {cita.fecha_hora}.\n\nSaludos,\nTu Equipo de Citas \n Laboratorio Dental - Sandra Gavíria',
                settings.DEFAULT_FROM_EMAIL,
                recipient_list,
                fail_silently=False,
            )

            service = get_google_calendar_service(request)
            if isinstance(service, HttpResponseRedirect):
                return service

            fecha_datetime = datetime.combine(fecha, hora)

            event = {
                'summary': 'Cita Programada',
                'location': 'Tu ubicación aquí',
                'description': f'Cita ID: {cita.id}\nCita programada para {cita.paciente.nombre}',
                'start': {
                    'dateTime': fecha_datetime.isoformat(),
                    'timeZone': 'America/Bogota',
                },
                'end': {
                    'dateTime': (fecha_datetime + timedelta(hours=1)).isoformat(),
                    'timeZone': 'America/Bogota',
                },
                'attendees': [
                    {'email': cita.paciente.correo},
                    {'email': "facturacionldsg@gmail.com"},
                ],
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'email', 'minutes': 1440},
                        {'method': 'popup', 'minutes': 10},
                    ],
                },
            }

            try:
                created_event = service.events().insert(calendarId='primary', body=event).execute()
                cita.google_event_id = created_event['id']
                cita.save()
                messages.success(request, 'Cita creada exitosamente y evento añadido a Google Calendar.')
            except HttpError as e:
                messages.warning(request, 'Cita creada exitosamente, pero hubo un problema al añadir el evento a Google Calendar.')
                print(f"Error al crear evento en Google Calendar: {str(e)}")
            except Exception as e:
                messages.error(request, f'Ocurrió un error: {str(e)}')

            return redirect('listcitas')
        else:
            for field, errors in formulario.errors.items():
                if field == '__all__':
                    for error in errors:
                        messages.error(request, f"Error general: {error}")
                else:
                    for error in errors:
                        messages.error(request, f"Error en el campo '{formulario.fields[field].label}': {error}")
            
            print("Errores en formulario:", formulario.errors)
    else:
        formulario = CitaForm(user=request.user)

    contexto = {
        'form': formulario,
        'titulo_formulario': 'Crear Cita',
        'is_superuser': request.user.is_superuser,
        'is_editing': False,
    }
    return render(request, 'crearcitas.html', contexto)


@never_cache
@login_required(login_url='acceso_denegado2')
def crearcitas2(request):
    return redirect('crearcitas')
    


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
@require_POST
def confirmar_actualizacion_cita(request, cita_id):
    try:
        cita = Cita.objects.get(id=cita_id)
        resultado = cita.confirmar_actualizacion()
        print("entró")
        if resultado:
            print("entró")

            service = get_google_calendar_service(request)
            if isinstance(service, HttpResponseRedirect):
                return service

            events_result = service.events().list(calendarId='primary', q=f'Cita ID: {cita_id}').execute()
            events = events_result.get('items', [])

            if events:
                event = events[0]
                service.events().delete(calendarId='primary', eventId=event['id']).execute()
                print(f"Evento de Google Calendar eliminado para la cita {cita_id}")
            
            print(f"Cita {cita_id} Completada.")
    
        if resultado:
            return JsonResponse({'status': 'success'}, status=200)
        else:
            return JsonResponse({'status': 'error', 'message': 'No se pudo actualizar la cita'}, status=400)
    except Cita.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Cita no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@never_cache
@login_required(login_url='acceso_denegado')
@require_GET
def get_horas_disponibles(request):
    fecha = request.GET.get('fecha')
    if fecha:
        try:
            fechas_disponibles = Fecha.objects.filter(fecha=fecha, disponible=True)
            citas_programadas = Cita.objects.filter(fecha_hora__fecha=fecha, estado='programada')
            horas_ocupadas = set(cita.fecha_hora.hora.strftime('%H:%M') for cita in citas_programadas)
            horas_disponibles = set(fecha_hora.hora.strftime('%H:%M') for fecha_hora in fechas_disponibles)
            horas_disponibles -= horas_ocupadas
            return JsonResponse(list(horas_disponibles), safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse([], safe=False)


@never_cache
@login_required(login_url='acceso_denegado')
def editarcitas(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id)
    fecha_hora_original = cita.fecha_hora

    if request.method == 'POST':
        form = CitaForm(request.POST, instance=cita, user=request.user)
        if form.is_valid():
            nueva_fecha = form.cleaned_data['fecha']
            nueva_hora = form.cleaned_data['hora']

            if not nueva_fecha or not nueva_hora:
                messages.error(request, 'Por favor, selecciona tanto la fecha como la hora para la cita.')
                return render(request, 'editarcitas.html', {'form': form, 'cita': cita, 'is_superuser': request.user.is_superuser, 'is_editing': True})

            nueva_cita = form.save(commit=False)
            nueva_fecha_hora, created = Fecha.objects.get_or_create(fecha=nueva_fecha, hora=nueva_hora)

            if nueva_fecha_hora != fecha_hora_original:
                citas_en_fecha_original = Cita.objects.filter(
                    fecha_hora=fecha_hora_original, 
                    estado__in=['programada', 'asistida']
                ).exclude(id=cita_id)
                
                if not citas_en_fecha_original.exists():
                    fecha_hora_original.disponible = True
                    fecha_hora_original.save()
                
                nueva_fecha_hora.disponible = False
                nueva_fecha_hora.save()
                
                recipient_list = [cita.paciente.correo, "facturacionldsg@gmail.com"]
                send_mail(
                    'Cita ReAgendada',
                    f'Hola {cita.paciente.nombre},\n\nTu cita para el {fecha_hora_original} ha sido reagendada para el {nueva_fecha_hora}.\n\n Saludos,\nTu Equipo de Citas \n Laboratorio Dental - Sandra Gavíria',
                    settings.DEFAULT_FROM_EMAIL,
                    recipient_list,
                    fail_silently=False,
                )

                nueva_cita.fecha_hora = nueva_fecha_hora

                service = get_google_calendar_service(request)
                if isinstance(service, HttpResponseRedirect):
                    return service

                if cita.google_event_id:
                    try:
                        service.events().delete(calendarId='primary', eventId=cita.google_event_id).execute()
                        print(f"Evento de Google Calendar eliminado para la cita {cita_id}")
                    except HttpError as e:
                        print(f"Error al eliminar el evento: {str(e)}")

            nueva_cita.save()

            fecha_datetime = datetime.combine(nueva_fecha, nueva_hora)
            event = {
                'summary': 'Cita Programada',
                'location': 'Tu ubicación aquí',
                'description': f'Cita ID: {nueva_cita.id}\nCita programada para {nueva_cita.paciente.nombre}',
                'start': {
                    'dateTime': fecha_datetime.isoformat(),
                    'timeZone': 'America/Bogota',
                },
                'end': {
                    'dateTime': (fecha_datetime + timedelta(hours=1)).isoformat(),
                    'timeZone': 'America/Bogota',
                },
                'attendees': [
                    {'correo': nueva_cita.paciente.correo},
                    {'correo': request.user.correo},
                ],
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'correo', 'minutes': 24 * 60},
                        {'method': 'popup', 'minutes': 10},
                    ],
                },
            }

            try:
                new_event = service.events().insert(calendarId='primary', body=event).execute()
                print('Nuevo evento creado: %s' % (new_event.get('htmlLink')))
                nueva_cita.google_event_id = new_event['id']
                nueva_cita.save()
                messages.success(request, 'Cita actualizada exitosamente y evento de Google Calendar actualizado.')
            except HttpError as e:
                print(f"Error al crear nuevo evento en Google Calendar: {str(e)}")
                messages.warning(request, 'Cita actualizada exitosamente, pero hubo un problema al añadir el evento a Google Calendar.')

            return redirect('listcitas')
        else:
            for field, errors in form.errors.items():
                if field == '__all__':
                    for error in errors:
                        messages.error(request, f"Error general: {error}")
                else:
                    for error in errors:
                        messages.error(request, f"Error en el campo '{form.fields[field].label}': {error}")
            
            print("Errores en formulario:", form.errors)
    else:
        initial_data = {
            'fecha': cita.fecha_hora.fecha,
            'hora': cita.fecha_hora.hora,
            'motivo': cita.motivo,
            'paciente': cita.paciente
        }
        form = CitaForm(instance=cita, user=request.user, initial=initial_data)

    contexto = {
        'form': form,
        'cita': cita,
        'is_superuser': request.user.is_superuser,
        'is_editing': True,
    }
    return render(request, 'editarcitas.html', contexto)


@never_cache
@login_required(login_url='acceso_denegado')
def citas_agendadas(request):
    
    ahora = timezone.now()
    if request.user.is_superuser:
        citas = Cita.objects.all()
    else:
        citas = Cita.objects.filter(paciente=request.user)

    for cita in citas:
        fecha_hora_cita = datetime.combine(cita.fecha_hora.fecha, cita.fecha_hora.hora)
        fecha_hora_cita = timezone.make_aware(fecha_hora_cita)
        if fecha_hora_cita < ahora and cita.estado == 'Programada':
            cita.estado = 'Inasistida'
            cita.save()

    fecha = request.GET.get('fecha')
    motivo = request.GET.get('motivo')
    estado = request.GET.get('estado')

    if fecha:
        citas = citas.filter(fecha_hora__fecha=fecha)
    if motivo:
        citas = citas.filter(motivo=motivo)
    if estado:
        citas = citas.filter(estado=estado)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        citas_data = [{
            'paciente': cita.paciente.documento,
            'nombres': cita.paciente.nombre,
            'fecha': cita.fecha_hora.fecha.strftime("%Y-%m-%d"),
            'hora': cita.fecha_hora.hora.strftime("%H:%M"),
            'motivo': cita.get_motivo_display(),
            'estado': cita.get_estado_display(),
            'id': cita.id,
            'editar_url': request.build_absolute_uri(f'/editarcitas/{cita.id}'),
            'is_superuser': request.user.is_superuser,
        } for cita in citas]

        return JsonResponse({'citas': citas_data})

    return render(request, 'listcitas.html', {
        'citas': citas,
        'motivos': [motivo[0] for motivo in Cita.MOTIVO_CHOICES],
        'estados': [estado[0] for estado in Cita.ESTADO_CHOICES],
    })


def reporte_citas_excel(request):
    data = json.loads(request.body)

    fecha_filtro = data.get('fecha', None)
    motivo_filtro = data.get('motivo', None)
    estado_filtro = data.get('estado', None)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    logger.debug(f"Fecha filtro: {fecha_filtro}")
    logger.debug(f"Motivo filtro: {motivo_filtro}")
    logger.debug(f"Estado filtro: {estado_filtro}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Citas Agendadas"

    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 21
    img.width = 39
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:F1')
    ws['B1'] = "LABORATORIO DENTAL"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Citas Agendadas"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Paciente", "Nombres", "Fecha", "Hora", "Motivo", "Estado"]
    ws.append(headers)

    header_fill = PatternFill(start_color="cab97d", end_color="cab97d", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    citas = Cita.objects.all()

    if fecha_filtro:
        citas = citas.filter(fecha_hora__fecha=fecha_filtro)

    if motivo_filtro:
        citas = citas.filter(motivo=motivo_filtro)

    if estado_filtro:
        citas = citas.filter(estado=estado_filtro)

    for cita in citas:
        paciente = str(cita.paciente)
        nombre = str(cita.paciente.nombre)
        motivo = cita.get_motivo_display()
        estado = cita.get_estado_display()

        if cita.fecha_hora:
            fecha = cita.fecha_hora.fecha.strftime('%d-%m-%Y') if hasattr(cita.fecha_hora.fecha, 'strftime') else str(
                cita.fecha_hora.fecha)
            hora = cita.fecha_hora.hora.strftime('%H:%M') if hasattr(cita.fecha_hora.hora, 'strftime') else str(
                cita.fecha_hora.hora)
        else:
            fecha = 'N/A'
            hora = 'N/A'

        ws.append([
            paciente,
            nombre,
            fecha,
            hora,
            motivo,
            estado
        ])

    column_widths = [20, 30, 15, 10, 30, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if cell.column_letter == 'C':
                cell.number_format = 'DD-MM-YYYY'
            elif cell.column_letter == 'D':
                cell.number_format = 'HH:MM'

    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_citas.xlsx"'

    wb.save(response)
    return response


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.logo_path = finders.find('img/logo.png')

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(letter[0] - 30, 30, f"Página {self._pageNumber} de {page_count}")
        
        self.saveState()
        img = ImageReader(self.logo_path)
        img_width = 4 * inch
        img_height = img_width * img.getSize()[1] / img.getSize()[0]
        
        x = (letter[0] - img_width) / 2
        y = (letter[1] - img_height) / 2
        
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.1))
        self.drawImage(img, x, y, width=img_width, height=img_height, mask='auto')
        self.restoreState()


def reporte_citas_pdf(request):
    data = json.loads(request.body)

    fecha_filtro = data.get('fecha', None)
    motivo_filtro = data.get('motivo', None)
    estado_filtro = data.get('estado', None)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    logger.debug(f"Fecha filtro: {fecha_filtro}")
    logger.debug(f"Motivo filtro: {motivo_filtro}")
    logger.debug(f"Estado filtro: {estado_filtro}")

    # Crear buffer para el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

    elements = []

    # Definir estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=1))

    # Títulos
    elements.append(Paragraph("LABORATORIO DENTAL", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Citas Agendadas", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Crear la tabla de datos
    data = [["Paciente", "Nombres", "Fecha", "Hora", "Motivo", "Estado"]]

    # Filtrar las citas según los parámetros
    citas = Cita.objects.all()

    if fecha_filtro:
        citas = citas.filter(fecha_hora__fecha=fecha_filtro)
    if motivo_filtro:
        citas = citas.filter(motivo=motivo_filtro)
    if estado_filtro:
        citas = citas.filter(estado=estado_filtro)

    # Agregar las citas a la tabla
    for cita in citas:
        paciente = str(cita.paciente)
        nombre = str(cita.paciente.nombre)
        motivo = cita.get_motivo_display()
        estado = cita.get_estado_display()

        if cita.fecha_hora:
            fecha = cita.fecha_hora.fecha.strftime('%d-%m-%Y') if hasattr(cita.fecha_hora.fecha, 'strftime') else str(
                cita.fecha_hora.fecha)
            hora = cita.fecha_hora.hora.strftime('%H:%M') if hasattr(cita.fecha_hora.hora, 'strftime') else str(
                cita.fecha_hora.hora)
        else:
            fecha = 'N/A'
            hora = 'N/A'

        # Agregar cada cita a los datos
        data.append([paciente, nombre, fecha, hora, motivo, estado])

    # Crear la tabla en PDF
    table = Table(data)

    # Estilo para la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cab97d")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    elements.append(table)

    # Generar el documento PDF
    doc.build(elements)

    # Preparar la respuesta para la descarga del archivo PDF
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_citas.pdf"'

    return response


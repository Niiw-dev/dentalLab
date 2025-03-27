import os
import sys

import openpyxl
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.staticfiles import finders
from django.core.files.storage import FileSystemStorage
from django.db import IntegrityError, transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.cache import never_cache

from google.auth.transport.requests import Request

from inicio.forms import FechaForm, UserForm
from inicio.models import Fecha, UserProfile

from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.colors import Color
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

import inicio.views as traer


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def crearfechas(request):
    if request.method == 'POST':
        form = FechaForm(request.POST)
        if form.is_valid():
            fecha = form.cleaned_data['fecha']
            hora = form.cleaned_data['hora']
            hora_final = form.cleaned_data['hora_final']

            if fecha == timezone.now().date():
                messages.error(request, 'No se pueden crear fechas para el día actual.')
                return render(request, 'crearfechas.html', {'form': form})

            current_time = timezone.datetime.combine(fecha, hora)
            end_time = timezone.datetime.combine(fecha, hora_final)

            try:
                with transaction.atomic():
                    while current_time <= end_time:
                        current_hour = current_time.time().hour
                        current_minute = current_time.time().minute

                        if (7 <= current_hour < 12) or (14 <= current_hour < 17) or (current_hour == 12 and current_minute == 0) or (current_hour == 17 and current_minute == 0):
                            if Fecha.objects.filter(fecha=current_time.date(), hora=current_time.time()).exists():
                                raise IntegrityError("Ya existe una fecha y hora para esa disponibilidad.")
                            
                            fecha_hora = Fecha(
                                fecha=current_time.date(),
                                hora=current_time.time(),
                            )
                            fecha_hora.save()

                        current_time += timezone.timedelta(minutes=20)

                messages.success(request, 'Fechas creadas con éxito.')
                return redirect('listfechas')
            except IntegrityError:
                messages.error(request, 'Ya existe una fecha y hora para esa disponibilidad.')
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = FechaForm()

    return render(request, 'crearfechas.html', {'form': form})

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def listfechas(request):
    fecha_actual = timezone.now().date()
    
    disponibilidades = Fecha.objects.filter(
        fecha__gte=fecha_actual,
        disponible=True
    ).order_by('fecha', 'hora')

    fecha_filtro = request.GET.get('fecha')
    if fecha_filtro:
        disponibilidades = disponibilidades.filter(fecha=fecha_filtro)

    context = {
        'disponibilidades': disponibilidades
    }

    return render(request, 'listfechas.html', context)


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def eliminarfechas(request, id):
    # print(f"request: {request}")
    # print(f"id: {id}")
    # sys.exit()
    fechaseliminar = get_object_or_404(Fecha, id=id)
    
    if request.method == 'POST':
        fechaseliminar.delete()
        return redirect('listfechas')
    
    return redirect('listfechas')


def reporte_fechas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fechas Agendadas"

    logo_path = finders.find('img/logo.png')

    img = Image(logo_path)
    img.height = 31
    img.width = 39
    ws.add_image(img, 'A1')

    ws.merge_cells('B1:B1')
    ws['B1'] = "LABORATORIO DENTAL"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:B2')
    ws['A2'] = "Fechas Agendadas"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Fecha", "Hora"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="cab97d", end_color="cab97d", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fechas = Fecha.objects.all()
    for fecha in fechas:
        fecha_hora = fecha.fecha
        hora = fecha.hora

        ws.append([
            fecha_hora.strftime('%d-%m-%Y'),
            hora.strftime('%H:%M'),
        ])

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 48

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if cell.column_letter == 'A':
                cell.number_format = 'DD-MM-YYYY'
            elif cell.column_letter == 'B':
                cell.number_format = 'HH:MM'
    
    for cell in ws[3]:
        cell.border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_fechas.xlsx"'
    
    wb.save(response)
    return response


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.logo_path = finders.find('img/logo.png')

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(letter[0] - 30, 30, f"Página {self._pageNumber} de {page_count}")
        
        self.saveState()
        img = ImageReader(self.logo_path)
        img_width = 4 * inch
        img_height = img_width * img.getSize()[1] / img.getSize()[0]
        
        x = (letter[0] - img_width) / 2
        y = (letter[1] - img_height) / 2
        
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.1))
        self.drawImage(img, x, y, width=img_width, height=img_height, mask='auto')
        self.restoreState()


def reporte_fechas_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=1))

    elements.append(Paragraph("LABORATORIO DENTAL", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Fechas Agendadas", styles['Heading2']))
    elements.append(Spacer(1, 12))

    data = [["Fecha", "Hora"]]

    fechas = Fecha.objects.all()
    for fecha in fechas:
        fecha_hora = fecha.fecha.strftime('%d-%m-%Y') if hasattr(fecha.fecha, 'strftime') else str(fecha.fecha)
        hora = fecha.hora.strftime('%H:%M') if hasattr(fecha.hora, 'strftime') else str(fecha.hora)
        data.append([fecha_hora, hora])

    table = Table(data)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cab97d")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    elements.append(table)

    doc.build(elements, canvasmaker=NumberedCanvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_fechas.pdf"'
    
    return response


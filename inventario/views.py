import json

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.cache import never_cache

from inicio.forms import InventarioForm
from inicio.models import Inventario
import inicio.views as traer

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

from inicio.forms import UserForm
from inicio.models import UserProfile
from inicio import views as traer
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
from google.auth.transport.requests import Request
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.colors import Color
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfgen import canvas

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.http import HttpResponse
from django.contrib.staticfiles import finders

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def eliminarelementos(request, id):
    elementoseliminar = get_object_or_404(Inventario, id=id)
    
    if request.method == 'POST':
        elementoseliminar.delete()
        return redirect('listelementos')
    
    return redirect('listelementos')

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def crearelementos(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('listelementos')  
    else:
        form = InventarioForm()
    
    return render(request, 'crearelementos.html', {'form': form})

@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def listelementos(request):
    # Obtener todos los elementos inicialmente
    inventarios = Inventario.objects.all()

    # Filtro por estado si se envía el parámetro desde el formulario
    estado_filtro = request.GET.get('estado')
    print(f"Estado filtro: {estado_filtro}")  # Para depurar el valor recibido

    if estado_filtro:  # Si hay un estado seleccionado
        inventarios = inventarios.filter(estado=estado_filtro)
    else:
        inventarios = Inventario.objects.all()  # Devuelve todos los elementos

    context = {
        'inventarios': inventarios,
        'estados': Inventario.ESTADO,  # Suponiendo que `ESTADO_CHOICES` está definido en el modelo
    }

    return render(request, 'listelementos.html', context)


@never_cache
@login_required(login_url='acceso_denegado')
@user_passes_test(traer.es_superusuario, login_url='acceso_denegado')
def editarelementos(request, id):
    form_edelem = Inventario.objects.get(id=id)
    if request.method == 'POST':
        form = InventarioForm(request.POST, request.FILES, instance=form_edelem)
        if form.is_valid():
            form.save()
            return redirect('listelementos') 
    else:
        form = InventarioForm(instance=form_edelem)
    return render(request, 'editarelementos.html', {'form': form})


def reporte_inventario_excel(request):
    # Obtener los filtros desde el cuerpo de la solicitud
    data = json.loads(request.body)
    filtroEstado = data.get('filtroEstados', None)  # Suponemos que 'filtroEstado' es uno de los filtros posibles

    # Crear un libro de trabajo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Insertar el logo
    logo_path = finders.find('img/logo.png')
    img = Image(logo_path)
    img.height = 38  # Ajusta el tamaño según se requiera
    img.width = 78
    ws.add_image(img, 'A1')

    # Títulos de la hoja
    ws.merge_cells('B1:C1')
    ws['B1'] = "LABORATORIO DENTAL"
    ws['B1'].font = Font(size=24, bold=True)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = "Inventario"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados de la tabla
    headers = ["Producto", "Cantidad", "Estado"]
    ws.append(headers)

    header_fill = PatternFill(start_color="cab97d", end_color="cab97d", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Estilo de bordes
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Filtrar el inventario según el filtro de estado
    inventarios = Inventario.objects.all()
    if filtroEstado:
        inventarios = inventarios.filter(estado=filtroEstado)  # Filtrando por estado del inventario

    # Rellenar las filas con los datos del inventario
    for inventario in inventarios:
        producto = inventario.producto or "N/A"
        cantidad = inventario.cantidad
        estado = dict(Inventario.ESTADO).get(inventario.estado, "N/A")

        ws.append([
            producto,
            cantidad,
            estado
        ])

    # Ajuste del ancho de las columnas
    column_widths = [25, 20, 40]  # Ajusta el ancho de las columnas
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Aplicar formato de alineación y bordes a las celdas
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    for cell in ws[3]:
        cell.border = border

    # Configurar la respuesta HTTP para descarga del archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_inventario.xlsx"'

    wb.save(response)
    return response


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []
        self.logo_path = finders.find('img/logo.png')

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(letter[0] - 30, 30, f"Página {self._pageNumber} de {page_count}")
        
        self.saveState()
        img = ImageReader(self.logo_path)
        img_width = 4 * inch
        img_height = img_width * img.getSize()[1] / img.getSize()[0]
        
        # Calculate position to center the image
        x = (letter[0] - img_width) / 2
        y = (letter[1] - img_height) / 2
        
        self.setFillColor(colors.Color(0, 0, 0, alpha=0.1))  # 10% opacity
        self.drawImage(img, x, y, width=img_width, height=img_height, mask='auto')
        self.restoreState()


def reporte_inventario_pdf(request):
    # Obtener los filtros desde el cuerpo de la solicitud
    data = json.loads(request.body)
    filtroEstado = data.get('filtroEstados', None)  # Filtro por estado

    # Crear el buffer y documento PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=1))

    elements.append(Paragraph("LABORATORIO DENTAL", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Inventario", styles['Heading2']))
    elements.append(Spacer(1, 12))

    # Cabecera de la tabla
    data = [["Producto", "Cantidad", "Estado"]]

    # Filtrar inventarios según el estado
    inventarios = Inventario.objects.all()
    if filtroEstado:
        inventarios = inventarios.filter(estado=filtroEstado)  # Filtrando por estado

    # Rellenar la tabla con los inventarios filtrados
    for inventario in inventarios:
        producto = inventario.producto or "N/A"
        cantidad = str(inventario.cantidad)
        estado = dict(Inventario.ESTADO).get(inventario.estado, "N/A")

        data.append([producto, cantidad, estado])

    # Crear la tabla con los datos
    table = Table(data, colWidths=[250, 100, 150])

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#cab97d")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])

    table.setStyle(style)

    # Agregar la tabla a los elementos del PDF
    elements.append(table)

    # Generar el PDF
    doc.build(elements)

    # Retornar el archivo PDF como respuesta
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_inventario.pdf"'

    return response

